#Importing necessary modules
from tkinter import *
from PIL import ImageTk,Image
import time
import tkinter.messagebox
import sqlite3 as s
from openpyxl import Workbook
import os
import os.path
#Database: Preset with these defaults, which aren't used directly in the program itself
students=['Andrew','Bob','Donald','Elisa','Jackson','Jake','James','Jamie','Jason','Jerry','John','Justin','Liam','Logan','Madison','Michael','Paul','Ronnie','Shawn','Tom']
ebooks=['Wuthering Heights',"Childhood's End",'Tale of Two Cities','Ragtime','Lord of the Flies','The Outsiders','A Wrinkle In Time','Romeo and Juliet','The Odyssey','To Kill a Mockingbird','Of Mice and Men','The Book Thief','Fahrenheit 451','Monster','The Giver','The Great Gatsby','Be More Chill','Great Expectations','The Lovely Bones','The Graveyard Book']
authors=['Emily Bronte','Arthur Clarke','Charles Dickens','E.L. Doctorow','William Golding','S.E. Hinton',"Madeleine L'Engle",'William Shakespeare','Homer','Harper Lee','John Steinbeck','Markus Zusak','Ray Bradbury','Walter Dean Myers','Lois Lowry','F. Scott Fitzgerald','Ned Vizzini','Charles Dickens','Alice Sebold','Neil Gaiman']
recodes=[8886359354, 2397314509, 6124958132, 2679585160, 5814797731, 4584899280, 8233295776, 9122594666, 4640588008, 9022443156, 432997300, 8769793297, 141970731, 3919872219, 3840422290, 7255088872, 544476650, 4048971721, 1313667766, 6854749152]
#Connecting to "Elib2.db": a SQL relational database used by this program
try:
    conn = s.connect("Elib2.db") #Trying to connect to db
except s.Error as e:
    print("Rerun the program please")
c=conn.cursor() #Creates cursor object to execute
try: #Checks if database has been committed already, so that the database doesn't create more rows every time the program is run
    c.execute("SELECT Committed FROM inventory WHERE Id=?",[1])
    committed=(c.fetchall())[0][0]
except s.Error:
    def create_table(create_table_sql):
        try:
            c.execute(create_table_sql)
        except s.Error as e:
            print(e)
    def add_values():
        school_table="""CREATE TABLE sampleclass1 (
                            Student text,
                            Id num,
                            Grade text,
                            Teacher text,
                            RedemptionCodes text,
                            Ebook text,
                            Author text,
                            Date of issue text
                        ); """
        inventory_table="""CREATE TABLE inventory (
                                Id num,
                                Ebooks text,
                                Authors text,
                                ReCodes text,
                                Committed text
                        ); """
        create_table(school_table)
        create_table(inventory_table)
        for item in range(0,len(students)):
            i=students[item]
            grade=9
            c.execute("INSERT INTO sampleclass1 (Student,Id,Grade,Teacher) VALUES (?,?,?,?)",[i,item,grade,teacher])
        for item in range(0,len(ebooks)):
            i=recodes[item]
            j=ebooks[item]
            k=authors[item]
            c.execute("INSERT INTO inventory VALUES (?,?,?,?,1)",[item,j,k,i])
        conn.commit()
    add_values()
weekly_report_factor=0
#Interface
tk=Tk()
tk.title("E-Lib")
#Book image
book_image=Image.open("bookbg.jpg") #Opens the image for python
book_image=book_image.resize((90,120),Image.ANTIALIAS) #Creates image size in pixels for the interface
book_image=ImageTk.PhotoImage(book_image) #Gets the image in tkinter format
#Q&A help image
help_button=Image.open("help_button.png")
help_button=help_button.resize((25,25),Image.ANTIALIAS)
help_button=ImageTk.PhotoImage(help_button)
canvas=Canvas(width=1000,height=750,bg="#ffe9d6")
canvas.pack()
def interface(week):
    #Data collection
    c.execute("SELECT Student FROM sampleclass1") #Selects data needed
    student_d=c.fetchall() #Fetches data selected in the line above from database in a list of tuples
    c.execute("SELECT Ebooks FROM inventory")
    ebook_d=c.fetchall()
    c.execute("SELECT Authors FROM inventory")
    authors_d=c.fetchall()
    c.execute("SELECT ReCodes FROM inventory")
    recode_d=c.fetchall()
    c.execute("SELECT Grade FROM sampleclass1")
    grade_d=c.fetchall()
    c.execute("SELECT Author FROM sampleclass1")
    author_d=c.fetchall()
    students3=[]
    #for-loops for setting data from tuple format to string
    for item in student_d:
        students3.append(item[0]) #Adds student names in string format from tuple format
    students=sorted(students3) #Sorts the names alphabetically
    ebooks1=[]
    for item in ebook_d:
        ebooks1.append(item[0]) #Adds ebook names in string format from tuple format
    ebooks=[]
    for ebook in ebooks1:
        if len(ebook)>15:
            loc=int((len(ebook)-(len(ebook)%2))/2)
            while ebook[loc]!=" ":
                loc=loc-1
            ebook=ebook[:loc]+'\n'+ebook[(loc+1):] #Adds a line break between titles for the program to use in E-Lib Catalog
        ebooks.append(ebook)
    authors1=[]
    authors=[]
    for item in authors_d: #Authors of inventory books
        authors1.append(item[0]) #Adds author names in string format
    for author in authors1:
        if len(author)>15:
            loc=int((len(author)-(len(author)%2))/2)
            while author[loc]!=" ":
                loc=loc-1
            author=author[:loc]+'\n'+author[(loc+1):] #Adds a line break between author names for use in E-Lib Catalog
        authors.append(author)
    recodes=[]
    for item in recode_d:
        recodes.append(item[0]) #Adds redemption codes from tuple format to string
    grades=[]
    for item in grade_d:
        grades.append(item[0])
    assigned_books_authors=[]
    for item in author_d: #Authors of books assigned
        assigned_books_authors.append(item[0])
    #Title and top
    canvas.create_text(500,50,text='E-Lib',font="Times 24")
    qa=canvas.create_image(570,50,image=help_button) #Gets the "help" image
    def qa_window(event): #Defines the Q&A window
        qa_rect=canvas.create_rectangle(10,10,990,740,fill="white")
        qa_title=canvas.create_text(490,50,font="Times 36",text="FAQ (Frequently Asked Questions):")
        qa_text="How am I supposed to assign books?\nIn the E-lib Catalog section, click on any of the books, and a window appears that tells you to select a student, and then assign the book, or cancel assigning. When you assign the book to a student, E-lib Log updates automatically and shows to whom your book is assigned.\n\nWhat if I want to unassign a book?\nE-Lib has one copy for each ebook, so you click on the book that is assigned already, and a window pops up which lets you \nunassign the book.\n\nHow do i edit a student or an ebook?\nHover your mouse over the 'Manage' dropdown menu, and it will give you 2 options: Either 'Students' or 'Ebooks' and then click on one of these buttons to edit what you want. If you click 'Students', then you can edit the student name and/or grade. If you click 'Ebooks', then you can edit the ebook name, redemption code, and/or author.\n\nDo I receive weekly reports? If so, how do I receive it?\nWeekly reports are available every Sunday. In fact, as soon as you run E-Lib on a Sunday, it would ask you whether you want to open your weekly report or not. Your weekly report is made with Microsoft Excel so that you can analyze detailed data."
        questionanswer=canvas.create_text(490,100,width=950,font="Times 14",text=qa_text,anchor=N,justify=CENTER)
        goback_button=canvas.create_rectangle(20,700,70,730,fill="yellow")
        goback_text=canvas.create_text(45,715,text="Back",font="Arial 12")
        def go_back(event):
            canvas.delete(qa_rect,qa_title,qa_text,questionanswer,goback_button,goback_text)
        canvas.tag_bind(goback_button,"<Button-1>",go_back)
        canvas.tag_bind(goback_text,"<Button-1>",go_back)    
    canvas.tag_bind(qa,"<Button-1>",qa_window)
    def show_help(event):
        help_t=canvas.create_text(570,70,text="(Help button)",font="Times 12")
        def hide_help(event):
            canvas.delete(help_t)
        canvas.tag_bind(qa,"<Leave>",hide_help)
    canvas.tag_bind(qa,"<Enter>",show_help)
    #E-lib Log
    canvas.create_rectangle(10,100,240,740,fill="#fceea9",outline="#dbdcdd")
    canvas.create_text(120,120,text="E-Lib Log:",font="Times 24",fill="black")
    now=time.localtime(time.time())
    date=time.strftime("%m/%d/%y",now)
    date="Date: "+date
    canvas.create_text(120,145,text=date,font="Times 16",fill="black")
    canvas.create_text(20,170,text="Students:  Grade  Books Assigned",font="Times 12",fill="black",anchor=W)
    posy=190
    for item in students:
        canvas.create_text(25,posy,text=item,anchor=W,font="Times 12",fill="black")
        posy=posy+25
    posy=190
    for item in range(0,len(grades)):
        c.execute("SELECT Grade FROM sampleclass1 WHERE Student=?",([students[item]]))
        student_g=c.fetchall()
        canvas.create_text(90,posy,text=student_g[0][0],anchor=W,font="Times 12",fill="black")
        posy=posy+25
    #Checking for any assignments to update E-Lib Log
    c.execute("SELECT Ebook FROM sampleclass1")
    assigned_books=c.fetchall()
    assigned_books1=[]
    for item in assigned_books:
        assigned_books1.append(item[0])
    for item in range(0,len(students)):
        id_factor=0
        while students3[id_factor]!=students[item]:
            id_factor=id_factor+1
        position=190+(25*item)
        if assigned_books1[id_factor]!=None:
            canvas.create_text(170,position,text=assigned_books1[id_factor],fill="Black",font="Times 12")
    #E-Lib Catalog
    canvas.create_rectangle(250,100,990,740,fill="#dee8d5",outline="brown")
    canvas.create_text(270,120,text="E-Lib Catalog:",anchor=W,font="Times 24",fill="black")
    px=270
    py=150
    def identify(event):
        #If-then-else sequences help the program know which title has been selected
        if event.x<570:
            if event.x<420:
                if event.y<450:
                    if event.y<300:
                        img_id=1
                    else:
                        img_id=2
                elif event.y<600:
                    img_id=3
                else:
                    img_id=4
            else:
                if event.y<450:
                    if event.y<300:
                        img_id=5
                    else:
                        img_id=6
                elif event.y<600:
                    img_id=7
                else:
                    img_id=8
        elif event.x>660:
            if event.x>810:
                if event.y<450:
                    if event.y<300:
                        img_id=17
                    else:
                        img_id=18
                elif event.y<600:
                    img_id=19
                else:
                    img_id=20
            else:
                if event.y<450:
                    if event.y<300:
                        img_id=13
                    else:
                        img_id=14
                elif event.y<600:
                    img_id=15
                else:
                    img_id=16
        else:
            if event.y<450:
                if event.y<300:
                    img_id=9
                else:
                    img_id=10
            elif event.y<600:
                img_id=11
            else:
                img_id=12
        #Action taken after book and title are known
        assign_factor=0
        for item in assigned_books1:
            if ebooks1[img_id-1]==item: #Checks if the book has not already been assigned    
                assign_factor=1
        if assign_factor!=1:
            #Assigning window/box
            book_window=canvas.create_rectangle(300,200,700,550,fill="white")
            book_title="Title of book: "+ebooks1[img_id-1]
            author_title="Author name: "+authors1[img_id-1]
            recode_title="Redemption code: "+recodes[img_id-1]
            bk_title=canvas.create_text(330,230,anchor=W,text=book_title,font="Times 16")
            bk_author=canvas.create_text(330,290,anchor=W,text=author_title,font="Times 16")
            bk_code=canvas.create_text(330,350,anchor=W,text=recode_title,font="Times 16")
            assignto=canvas.create_text(330,410,anchor=W,text="Assign to: ",font="Times 16")
            var=StringVar()
            var.set("Select a student:")
            student_assign=OptionMenu(tk,var,*students)
            student_assign1=canvas.create_window(500,410,window=student_assign)
            exit_assign=canvas.create_rectangle(640,500,690,540,fill="yellow")
            exit_assign1=canvas.create_text(665,520,text="Done")
            #Assign function made to finish assigning a student
            def assign(event):
                assignment=var.get() #Gets the selected option from the dropdown menu
                canvas.delete('all')
                manage=canvas.create_rectangle(850,25,950,50,fill="#a8a8a8")
                canvas.create_text(900,37,text="Manage")
                canvas.tag_bind(manage,"<Enter>",dropdown)
                canvas.tag_bind(manage,"<Leave>",dropdowno)
                c.execute("UPDATE sampleclass1 SET Ebook=? WHERE Student=?",(ebooks1[img_id-1],assignment))
                c.execute("UPDATE sampleclass1 SET RedemptionCodes=? WHERE Student=?",(recodes[img_id-1],assignment))
                c.execute("UPDATE sampleclass1 SET Author=? WHERE Student=?",(authors1[img_id-1],assignment))
                conn.commit() #Saves the changes made to the database
                interface(week) #Updates interface to new results dynamically
                tkinter.messagebox.showinfo("E-Lib Notification Center","\"%s\" has been assigned to %s"%(ebooks1[img_id-1],assignment))
            canvas.tag_bind(exit_assign,"<Button-1>",assign)
            canvas.tag_bind(exit_assign1,"<Button-1>",assign)
            def cancel1(event):
                #Deletes the window
                canvas.delete(book_window,exit_assign,exit_assign1,student_assign1,bk_title,bk_author,bk_code,assignto,cancel_button2,cancel_button3)
            #Cancel button, text, and their event bindings
            cancel_button2=canvas.create_rectangle(310,500,390,540,fill="#dee8d5")
            cancel_button3=canvas.create_text(350,520,text="Cancel",font="Times 12")
            canvas.tag_bind(cancel_button2,"<Button-1>",cancel1)
            canvas.tag_bind(cancel_button3,"<Button-1>",cancel1)
        else: #Creates a window for unassigning if the book is already assigned
            book_window=canvas.create_rectangle(300,200,700,550,fill="white")
            cancel_message=canvas.create_text(330,230,anchor=W,text="This book has already been assigned.\nPlease unassign this book or cancel",font="Times 16")
            unassign_button=canvas.create_rectangle(450,300,550,350,fill="yellow")
            unassign_2=canvas.create_text(500,325,text="Unassign",font="Times 12")
            def cancel(event):
                canvas.delete(book_window,cancel_message,unassign_button,unassign_2,cancel_button,cancel_button1)
            cancel_button=canvas.create_rectangle(310,500,390,540,fill="#dee8d5")
            cancel_button1=canvas.create_text(350,520,text="Cancel",font="Times 12")
            canvas.tag_bind(cancel_button,"<Button-1>",cancel)
            canvas.tag_bind(cancel_button1,"<Button-1>",cancel)
            #Unassign function
            def unassign(event):
                for item in range(0,len(assigned_books1)):
                    if assigned_books1[item]==ebooks1[img_id-1]:
                        #Cursor object executes command for unassigning
                        c.execute("UPDATE sampleclass1 SET Ebook=NULL WHERE Id=?",[item])
                        c.execute("UPDATE sampleclass1 SET RedemptionCodes=NULL WHERE Id=?",[item])
                        c.execute("UPDATE sampleclass1 SET Author=NULL WHERE Id=?",[item])
                        conn.commit() #Saves the transaction/change made to the database
                        interface(week) #Updates the interface so that results update dynamically without any refresh required
                        tkinter.messagebox.showinfo("E-Lib Notification Center","\"%s\" has been unassigned"%(ebooks1[img_id-1]))
            canvas.tag_bind(unassign_button,"<Button-1>",unassign)
            canvas.tag_bind(unassign_2,"<Button-1>",unassign)       
    aaa=0
    px=270
    py=150
    for x in range(0,5):
        for y in range(0,4):
            author="By:\n"+authors[aaa]
            zz=canvas.create_image(px,py,image=book_image,anchor=NW)
            yyy=canvas.create_text(px+45,py+20,text=ebooks[aaa],justify=CENTER,font="Times 9")
            zzz=canvas.create_text(px+45,py+90,text=author,justify=CENTER,font="Times 9")
            canvas.tag_bind(zz,"<Button-1>",identify)
            canvas.tag_bind(yyy,"<Button-1>",identify)
            canvas.tag_bind(zzz,"<Button-1>",identify)
            py=py+150
            aaa=aaa+1
        px=px+150
        py=150
    #Dropdown menu under "Manage" rectangle on top-right corner
    def dropdowno(event):
        if event.x<850 or event.x>950 or event.y<25 or event.y>100:
            canvas.delete('all')
            interface(week)
            manage=canvas.create_rectangle(850,25,950,50,fill="#a8a8a8")
            canvas.tag_bind(manage,"<Enter>",dropdown)
            canvas.tag_bind(manage,"<Leave>",dropdowno)
            canvas.create_text(900,37,text="Manage")
    def dropdown(event):
        #Window for editing student information
        def student_window(event):
            canvas.create_rectangle(20,20,980,730,fill="#dbe8ff",outline="#dbe8ff")
            canvas.create_text(500,50,text="Students List Editbox",font="Times 36")
            def edit_student(event):
                if var1.get()=="Select a student:":
                    tkinter.messagebox.showerror("E-Lib Notification Center","Please select a student name")
                else:
                    student_name="Student name: "+var1.get()
                    studentname=canvas.create_text(300,300,anchor=W,text=student_name,font="Times 14")
                    student=[var1.get()]
                    c.execute("SELECT Grade FROM sampleclass1 WHERE Student=?",(student)) #TRY TO FIND GRADE
                    grade=c.fetchall()
                    id_factor=0
                    for item in range(0,len(students3)):
                        if students3[item]==var1.get():
                            id1=id_factor
                        else:
                            id_factor=id_factor+1
                    grade="Grade: "+grades[id1]
                    Grade=canvas.create_text(300,330,anchor=W,text=grade,font="Times 14")
                    editname=canvas.create_text(350,360,anchor=W,text="Edit name:",font="Times 14")
                    def write_name():
                        name_write=Entry(relief=FLAT)
                        name_write1=canvas.create_window(520,360,window=name_write)
                        canvas.delete(edit_name1)
                        click_factor=0
                        def done():
                            click_factor=1
                            if name_write.get()=="" or (name_write.get()).isspace()==True:
                                tkinter.messagebox.showerror("E-Lib Notification Center","Please type a name")
                            elif name_write.get()==var1.get():
                                tkinter.messagebox.showerror("E-Lib Notification Center","This name is the original name. Please use a newer name.")
                            else:
                                c.execute("UPDATE sampleclass1 SET Student=? WHERE Student=?",(name_write.get(),var1.get()))
                                c.execute("SELECT Student FROM sampleclass1")
                                tkinter.messagebox.showinfo("E-Lib Notification Center","Student name has been changed from %s to %s"%(var1.get(),name_write.get()))
                        done=Button(text="Done",command=done)
                        canvas.create_window(650,360,window=done)
                    edit_name=Button(text="Edit",command=write_name)
                    edit_name1=canvas.create_window(450,360,window=edit_name)
                    editgrade=canvas.create_text(350,390,anchor=W,text="Edit grade:",font="Times 14")
                    def write_grade():
                        grade_write=Entry(relief=FLAT)
                        canvas.create_window(520,390,window=grade_write)
                        canvas.delete(edit_grade1)
                        def done1():
                            if grade_write.get()=="" or (grade_write.get()).isspace()==True:
                                tkinter.messagebox.showerror("E-Lib Notification Center","Please type a grade")
                            else:
                                id_factor=0
                                for item in range(0,len(students3)):
                                    if students3[item]==var1.get():
                                        id1=id_factor
                                    else:
                                        id_factor=id_factor+1
                                c.execute("UPDATE sampleclass1 SET Grade=? WHERE Id=?",(grade_write.get(),id1))
                                tkinter.messagebox.showinfo("E-Lib Notification Center","Student grade has been changed to %s"%(grade_write.get()))
                        done1=Button(text="Done",command=done1)
                        canvas.create_window(650,390,window=done1)
                    edit_grade=Button(text="Edit",command=write_grade)
                    edit_grade1=canvas.create_window(450,390,window=edit_grade)
            canvas.create_text(50,100,text="Instructions:\nSelect a student name to edit his name and/or grade:",anchor=NW,font="Times 16")
            var1=StringVar()
            var1.set("Select a student:")
            edit_select_student=OptionMenu(tk,var1,*students)
            canvas.create_window(700,135,window=edit_select_student)
            edit_button=canvas.create_rectangle(780,120,830,150,fill="#9172ff",outline="#9172ff")
            edit_text=canvas.create_text(805,135,text="Edit")
            canvas.tag_bind(edit_button,"<Button-1>",edit_student)
            canvas.tag_bind(edit_text,"<Button-1>",edit_student)
            #Defines cancel function
            def cancel(event):
                canvas.delete('all')
                interface(week)
            def done(event):
                conn.commit()
                canvas.delete('all')
                interface(week)
            cancel_button4=canvas.create_rectangle(30,680,80,710,fill="yellow",outline="yellow")
            cancel_text4=canvas.create_text(55,695,text="Cancel")
            done_button=canvas.create_rectangle(920,680,970,710,fill="yellow",outline="yellow")
            done_text=canvas.create_text(945,695,text="Done")
            canvas.tag_bind(cancel_button4,"<Button-1>",cancel)
            canvas.tag_bind(cancel_text4,"<Button-1>",cancel)
            canvas.tag_bind(done_button,"<Button-1>",done)
            canvas.tag_bind(done_text,"<Button-1>",done)
        def ebook_window(event):
            #Window for editing ebook information
            canvas.create_rectangle(20,20,980,730,fill="#ebedbd",outline="#ebedbd")
            canvas.create_text(500,50,text="Ebooks List Editbox",font="Times 36")
            def edit_ebook(event):
                if var2.get()=="Select an ebook:":
                    tkinter.messagebox.showerror("E-Lib Notification Center","Please select a student name")
                else:
                    ebook_name="Ebook name: "+var2.get()
                    ebookname=canvas.create_text(300,300,anchor=W,text=ebook_name,font="Times 14")
                    id_factor=0
                    while ebooks1[id_factor]!=var2.get():
                        id_factor=id_factor+1
                    recode="Redemption Code: "+recodes[id_factor]
                    author="Author Name: "+authors1[id_factor]
                    canvas.create_text(300,330,anchor=W,text=recode,font="Times 14")
                    canvas.create_text(300,360,anchor=W,text=author,font="Times 14")
                    canvas.create_text(350,390,anchor=W,text="Edit name:",font="Times 14")
                    def write_name():
                        name_write=Entry(relief=FLAT)
                        name_write1=canvas.create_window(520,390,window=name_write)
                        canvas.delete(edit_name1)
                        click_factor=0
                        def done():
                            click_factor=1
                            if name_write.get()=="" or (name_write.get()).isspace()==True:
                                tkinter.messagebox.showerror("E-Lib Notification Center","Please type a name")
                            elif name_write.get()==var2.get():
                                tkinter.messagebox.showerror("E-Lib Notification Center","This name is already used by the ebook. Please use a newer name.")
                            else:
                                id_factor=0
                                while ebooks1[id_factor]!=var2.get():
                                    id_factor=id_factor+1
                                c.execute("UPDATE inventory SET Ebooks=? WHERE Ebooks=?",(name_write.get(),var2.get()))
                                c.execute("UPDATE sampleclass1 SET Ebook=? WHERE Ebook=?",(name_write.get(),var2.get()))
                                tkinter.messagebox.showinfo("E-Lib Notification Center","Ebook name has been changed from %s to %s"%(var2.get(),name_write.get()))
                        done=Button(text="Done",command=done)
                        canvas.create_window(650,390,window=done)
                    edit_name=Button(text="Edit",command=write_name)
                    edit_name1=canvas.create_window(450,390,window=edit_name)
                    editgrade=canvas.create_text(350,420,anchor=W,text="Edit code:",font="Times 14")
                    def write_code():
                        code_write=Entry(relief=FLAT)
                        canvas.create_window(520,420,window=code_write)
                        canvas.delete(edit_code1)
                        def done1():
                            if code_write.get()=="" or (code_write.get()).isspace()==True:
                                tkinter.messagebox.showerror("E-Lib Notification Center","Please type a code")
                            else:
                                id_factor=0
                                while ebooks1[id_factor]!=var2.get():
                                    id_factor=id_factor+1
                                c.execute("UPDATE inventory SET ReCodes=? WHERE Id=?",(code_write.get(),id_factor))
                                tkinter.messagebox.showinfo("E-Lib Notification Center","Ebook code has been changed to %s"%(code_write.get()))
                        done1=Button(text="Done",command=done1)
                        canvas.create_window(650,420,window=done1)
                    edit_code=Button(text="Edit",command=write_code)
                    edit_code1=canvas.create_window(450,420,window=edit_code)
                    def write_author():
                        author_write=Entry(relief=FLAT)
                        canvas.create_window(520,450,window=author_write)
                        canvas.delete(edit_author1)
                        def done2():
                            if author_write.get()=="" or (author_write.get()).isspace()==True:
                                tkinter.messagebox.showerror("E-Lib Notification Center","Please type a name")
                            else:
                                id_factor=0
                                while ebooks1[id_factor]!=var2.get():
                                    id_factor=id_factor+1
                                c.execute("UPDATE inventory SET Authors=? WHERE Id=?",(author_write.get(),id_factor))
                                tkinter.messagebox.showinfo("E-Lib Notification Center","Author name has been changed to %s"%(author_write.get()))
                        done2=Button(text="Done",command=done2)
                        canvas.create_window(650,450,window=done2)
                    canvas.create_text(350,450,anchor=W,text="Edit author:",font="Times 14")
                    edit_author=Button(text="Edit",command=write_author)
                    edit_author1=canvas.create_window(460,450,window=edit_author)
            canvas.create_text(50,100,text="Instructions:\nSelect an ebook name to edit its name, author, and/or redemption code:",anchor=W,font="Times 16")
            var2=StringVar()
            var2.set("Select an ebook:")
            edit_select_ebook=OptionMenu(tk,var2,*ebooks1)
            canvas.create_window(750,110,window=edit_select_ebook)
            edit_button=canvas.create_rectangle(850,95,900,125,fill="yellow",outline="yellow")
            edit_text=canvas.create_text(875,110,text="Edit")
            canvas.tag_bind(edit_button,"<Button-1>",edit_ebook)
            canvas.tag_bind(edit_text,"<Button-1>",edit_ebook)
            #Defines cancel function
            def cancel(event):
                canvas.delete('all')
                interface(week)
            def done(event):
                conn.commit()
                canvas.delete('all')
                interface(week)
            cancel_button4=canvas.create_rectangle(30,680,80,710,fill="#ef977c",outline="#ef977c")
            cancel_text4=canvas.create_text(55,695,text="Cancel")
            done_button=canvas.create_rectangle(920,680,970,710,fill="#ef977c",outline="#ef977c")
            done_text=canvas.create_text(945,695,text="Done")
            canvas.tag_bind(cancel_button4,"<Button-1>",cancel)
            canvas.tag_bind(cancel_text4,"<Button-1>",cancel)
            canvas.tag_bind(done_button,"<Button-1>",done)
            canvas.tag_bind(done_text,"<Button-1>",done)
        #Dropdown menu rectangles
        Students=canvas.create_rectangle(850,50,950,75,fill="#e0e0e0")
        Students1=canvas.create_text(900,62,text="Students")
        Ebooks=canvas.create_rectangle(850,75,950,100,fill="#e0e0e0")
        Ebooks1=canvas.create_text(900,87,text="Ebooks")
        canvas.tag_bind(Students,"<Button-1>",student_window)
        canvas.tag_bind(Students1,"<Button-1>",student_window)
        canvas.tag_bind(Students,"<Leave>",dropdowno)
        canvas.tag_bind(Ebooks,"<Button-1>",ebook_window)
        canvas.tag_bind(Ebooks1,"<Button-1>",ebook_window)
        canvas.tag_bind(Ebooks,"<Leave>",dropdowno)
    #"Manage" rectangle for dropdown menu
    manage=canvas.create_rectangle(850,25,950,50,fill="#a8a8a8")
    canvas.create_text(900,37,text="Manage")
    canvas.tag_bind(manage,"<Enter>",dropdown)
    canvas.tag_bind(manage,"<Leave>",dropdowno)
    #Weekly Report Generator
    if time.strftime("%a")=="Wed":
        wb=Workbook()
        ws=wb.active
        ws['A1']='E-Lib Stats: Week of %s'%(time.strftime("%b %d"))
        ws['B1']='Student Name'
        ws['C1']='Grade'
        ws['D1']='Book Issued'
        ws['E1']='Book Redemption Code'
        ws['F1']='Book Author'
        for item in range(0,len(students)):
            cell='B'+str(item+2)
            ws[cell]=students[item]
            id_factor=0
            for item1 in range(0,len(students3)):
                if students3[item1]==students[item]:
                    id1=id_factor
                else:
                    id_factor=id_factor+1
            grade_cell='C'+str(item+2)
            ws[grade_cell]=int(grades[id1])
        for item in range(0,len(assigned_books1)):
            cell='D'+str(item+2)
            ws[cell]=assigned_books1[item]
        for item in range(0,len(recodes)):
            cell='E'+str(item+2)
            ws[cell]=recodes[item]
        for item in range(0,len(assigned_books_authors)):
            cell='F'+str(item+2)
            ws[cell]=assigned_books_authors[item]
        filename="E-Lib_Report_%s.xlsx"%(time.strftime("%b %d"))
        wb.save(filename)
        if week==0:
            weekly_report_window=canvas.create_rectangle(200,250,800,500,fill="#e5efff",outline="#e5efff")
            weekly_report_text=canvas.create_text(500,260,anchor=N,justify=CENTER,text='Happy Sunday!\nE-Lib provides administrators with weekly reports\nevery Sunday to show whom books are assigned! Would you\nlike to view the report as an Excel file? *Note: Please close the excel\nfile before you close and rerun this program.',font="Times 16")
            def approve():
                filename="E-Lib_Report_%s.xlsx"%(time.strftime("%b %d"))
                os.startfile(filename)
                canvas.delete(weekly_report_window,weekly_report_text,approve_button1,disapprove_button1)
            def disapprove():
                canvas.delete(weekly_report_window,weekly_report_text,approve_button1,disapprove_button1)
            approve_button=Button(text="Yes",command=approve)
            approve_button1=canvas.create_window(600,470,window=approve_button)
            disapprove_button=Button(text="No",command=disapprove)
            disapprove_button1=canvas.create_window(400,470,window=disapprove_button)
            week=1
interface(weekly_report_factor)#Starts the interface function, using var weekly_report_factor, which is used in finding out when the program is supposed to give its weekly reports
